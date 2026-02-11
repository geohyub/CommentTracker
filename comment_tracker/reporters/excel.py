"""Excel 다중 탭 리포트 생성."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..db import get_connection, get_db_info
from ..analytics.project_stats import get_all_projects_summary
from ..analytics.client_stats import get_all_clients_summary
from ..analytics.trend import get_category_trend_by_period
from ..analytics.recurring import find_recurring_themes


HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
GOOD_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
WARN_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
BAD_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)


def generate_stats_report(output_path, client=None, db_path=None):
    """5개 탭의 Excel 통계 리포트 생성."""
    wb = Workbook()

    # 탭 1: 개요
    ws1 = wb.active
    ws1.title = "개요"
    info = get_db_info(db_path)

    ws1.append(["Comment Tracker 통계 리포트"])
    ws1.merge_cells("A1:B1")
    ws1.cell(1, 1).font = Font(name="Calibri", bold=True, size=16)

    ws1.append([])
    ws1.append(["항목", "값"])
    style_header_row(ws1, 3, 2)

    ws1.append(["전체 프로젝트", info["project_count"]])
    ws1.append(["전체 배치", info["batch_count"]])
    ws1.append(["전체 코멘트", info["comment_count"]])
    ws1.append(["L&L 플래그", info["ll_flag_count"]])
    ws1.append(["기간", f"{info['date_from'] or 'N/A'} ~ {info['date_to'] or 'N/A'}"])
    ws1.append(["클라이언트", ", ".join(info["clients"]) if info["clients"] else "없음"])

    for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
    auto_width(ws1)

    # 탭 2: 프로젝트별
    ws2 = wb.create_sheet("프로젝트별")
    projects = get_all_projects_summary(db_path)
    proj_headers = ["프로젝트", "클라이언트", "유형", "코멘트 종류", "리비전수", "전체", "Major", "Minor", "감소율"]
    ws2.append(proj_headers)
    style_header_row(ws2, 1, len(proj_headers))

    for p in projects:
        reduction = f"{p['reduction']}%" if p.get("reduction") is not None else "N/A"
        comment_types = ", ".join(p.get("comment_types", [])) if p.get("comment_types") else "-"
        ws2.append([
            p["project_code"], p["client"], p.get("report_type", ""),
            comment_types, p["batch_count"], p["total_comments"],
            p["major_count"], p["minor_count"], reduction
        ])

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(proj_headers)):
        for cell in row:
            cell.border = THIN_BORDER
        red_cell = row[len(proj_headers) - 1]
        try:
            val = int(str(red_cell.value).replace("%", "").replace("N/A", "-1"))
            if val >= 70:
                red_cell.fill = GOOD_FILL
            elif val >= 40:
                red_cell.fill = WARN_FILL
            elif val >= 0:
                red_cell.fill = BAD_FILL
        except ValueError:
            pass
    auto_width(ws2)

    # 탭 3: 클라이언트별
    ws3 = wb.create_sheet("클라이언트별")
    clients = get_all_clients_summary(db_path)
    client_headers = ["클라이언트", "프로젝트수", "전체 코멘트", "Major", "Minor", "프로젝트당 평균"]
    ws3.append(client_headers)
    style_header_row(ws3, 1, len(client_headers))

    for c in clients:
        avg = round(c["total_comments"] / c["project_count"], 1) if c["project_count"] > 0 else 0
        ws3.append([
            c["client"], c["project_count"], c["total_comments"],
            c["major"], c["minor"], avg
        ])

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(client_headers)):
        for cell in row:
            cell.border = THIN_BORDER
    auto_width(ws3)

    # 탭 4: 카테고리 트렌드
    ws4 = wb.create_sheet("카테고리 트렌드")
    trend_data = get_category_trend_by_period(client=client, db_path=db_path)
    trend_headers = ["기간", "오타/문법", "가독성", "그림/표", "서식/용어", "참조/목차", "Minor 합계"]
    ws4.append(trend_headers)
    style_header_row(ws4, 1, len(trend_headers))

    for t in trend_data:
        ws4.append([
            t["period"], t["Typo"], t["Readability"],
            t["FigTable"], t["Format"], t["Reference"], t["total"]
        ])

    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=len(trend_headers)):
        for cell in row:
            cell.border = THIN_BORDER
    auto_width(ws4)

    # 탭 5: 반복 테마
    ws5 = wb.create_sheet("반복 테마")
    themes = find_recurring_themes(db_path=db_path)
    theme_headers = ["테마", "발생횟수", "프로젝트", "클라이언트", "카테고리"]
    ws5.append(theme_headers)
    style_header_row(ws5, 1, len(theme_headers))

    for t in themes[:30]:
        ws5.append([
            t["term"], t["occurrences"],
            ", ".join(t["projects"]),
            ", ".join(t["clients"]),
            t["primary_category"],
        ])

    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row, max_col=len(theme_headers)):
        for cell in row:
            cell.border = THIN_BORDER
    auto_width(ws5)

    wb.save(output_path)
    return output_path
