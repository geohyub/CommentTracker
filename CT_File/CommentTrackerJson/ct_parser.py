"""
Comment Tracker Batch Parser & Classifier
Parses EDF, Orsted, Vena format comment sheets and generates CT JSON.
"""
import openpyxl
import json
import re
import os
from datetime import datetime


# ============================================================
# Classification Engine
# ============================================================

EXCLUSION_PATTERNS = [
    r'when will .* be submitted',
    r'submission (schedule|timeline|date)',
    r'please provide (updated )?timeline',
    r'update distribution list',
    r'well written',
    r'good (job|work|report)',
    r'no (further )?comment',
    r'is .* included in (this|the) (report|scope)',
    r'out of scope',
    r'additional work beyond',
]

MAJOR_PATTERNS = [
    r'coordinate system',r'crs',r'datum',r'projection',r'utm zone',r'epsg',r'wkid',
    r'velocity model',r'sound velocity',r'wrong (correction|value)',
    r'misclassif',r'wrong depth',r'incorrect stratigraphy',
    r'wrong filter',r'wrong gain',r'incorrect (configuration|correction)',
    r'survey (boundary|extent|area)',r'wrong (survey|lot|area)',
    r'safety',r'hazard',r'risk',
    r'reprocess',r'resubmit data',
    r'not acceptable',r'critical',r'major',
    r'wrong sensor',r'incorrect (equipment|sensor)',
]

TYPO_PATTERNS = [
    r'typo',r'spelling',r'grammar',r'punctuation',r'misspell',
    r'should be ["\']?\w+["\']?',r'wrong (word|date|name)',r'rename',
    r'incorrect (spelling|name)',
]

FIGTABLE_PATTERNS = [
    r'\bfigure\b',r'\bfig[\.\s]',r'\btable\b',r'\bcaption\b',
    r'resolution',r'\bdpi\b',r'\blegend\b',r'\baxis\b',r'\bchart\b',
    r'\bmap\b',r'\bimage\b',r'scale bar',r'color palette',r'\bdiagram\b',
    r'\bphoto\b',r'screenshot',r'gantt',r'attachment',r'line plan',
    r'appendix.*visual',r'appendix.*figure',r'appendix.*map',
]

REFERENCE_PATTERNS = [
    r'\btoc\b',r'table of contents',r'cross.?reference',r'page number',
    r'\?\?',r'broken link',r'section reference',r'bibliography',
    r'\bcitation\b',r'document number',r'standard reference',
    r'appendix.*referenced.*not included',
]

FORMAT_PATTERNS = [
    r'consisten',r'terminolog',r'\bstyle\b',r'numbering',r'\bfont\b',
    r'formatting',r'\btemplate\b',r'\bheader\b',r'\bfooter\b',
    r'\bmargin\b',r'spacing',r'alignment',r'abbreviation',
    r'unit consisten',r'company name',r'cover (sheet|page)',r'title page',
    r'file name',r'naming convention',r'outdated',r'update.*logo',
    r'update.*name',r'update.*date',r'update.*version',
]

def classify_comment(text, priority_from_source=None):
    """Classify a comment. Returns (severity, category, confidence, excluded, exclude_reason)."""
    text_lower = text.lower().strip()

    # Step 1: Exclusion check
    for pat in EXCLUSION_PATTERNS:
        if re.search(pat, text_lower):
            return ("Minor", "Readability", "High", True, _exclusion_reason(text_lower))

    # Step 2: Major check
    if priority_from_source and str(priority_from_source).strip() in ('1', '2', 'Major', 'Critical'):
        return ("Major", "Technical", "High", False, None)

    for pat in MAJOR_PATTERNS:
        if re.search(pat, text_lower):
            return ("Major", "Technical", "High", False, None)

    # Step 3: Minor subcategories (priority order)
    for pat in TYPO_PATTERNS:
        if re.search(pat, text_lower):
            return ("Minor", "Typo", "High", False, None)
    for pat in FIGTABLE_PATTERNS:
        if re.search(pat, text_lower):
            return ("Minor", "FigTable", "High", False, None)
    for pat in REFERENCE_PATTERNS:
        if re.search(pat, text_lower):
            return ("Minor", "Reference", "High", False, None)
    for pat in FORMAT_PATTERNS:
        if re.search(pat, text_lower):
            return ("Minor", "Format", "High", False, None)

    return ("Minor", "Readability", "Medium", False, None)


def _exclusion_reason(text):
    if 'submit' in text or 'timeline' in text or 'schedule' in text:
        return "Schedule/timeline inquiry"
    if 'scope' in text:
        return "Scope inquiry"
    if 'distribution' in text:
        return "Administrative request"
    if any(w in text for w in ['well written','good job','good work','no comment','no further']):
        return "Positive feedback / no comment"
    return "Non-technical inquiry"


def extract_tags(text, comment_num):
    """Extract technical tags from comment text."""
    text_lower = text.lower()
    tags = [f"ref:{comment_num:02d}"]

    tag_map = {
        'resolution': [r'resolution', r'\bdpi\b'],
        'coordinate system': [r'coordinate', r'\bcrs\b', r'\bdatum\b', r'projection', r'\bepsg\b', r'\butm\b', r'\bwkid\b'],
        'velocity': [r'velocity', r'sound speed'],
        'bathymetry': [r'bathymetr', r'\bdepth\b', r'\bmbes\b', r'\bdtm\b'],
        'magnetic': [r'magnetom', r'\bmag\b'],
        'seismic': [r'\bsbp\b', r'\buhr\b', r'seismic', r'sub.?bottom'],
        'side scan': [r'side.?scan', r'\bsss\b'],
        'seabed': [r'seabed', r'sea.?floor'],
        'terminology': [r'terminolog', r'consisten.*term', r'abbreviat'],
        'positioning': [r'\bgnss\b', r'\busbl\b', r'navigation', r'position'],
        'cable route': [r'cable route', r'cable'],
        'sediment': [r'sediment', r'geolog'],
        'calibration': [r'calibrat'],
        'equipment': [r'vessel', r'equipment', r'sensor'],
        'figure': [r'\bfigure\b', r'\bfig[\.\s]'],
        'table': [r'\btable\b'],
        'tidal': [r'tidal', r'\btide\b'],
        'metadata': [r'metadata', r'attribute'],
        'GIS': [r'\bgis\b', r'\bshp\b', r'shape.?file', r'\btif\b'],
        'SVP': [r'\bsvp\b', r'sound velocity profile'],
    }

    for tag, patterns in tag_map.items():
        for pat in patterns:
            if re.search(pat, text_lower):
                tags.append(tag)
                break

    return ','.join(tags)


def generate_summary_ko(text, category):
    """Generate a Korean summary from English comment text."""
    text = text.strip()
    if '\n' in text:
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        first = lines[0]
        if len(first) > 200:
            first = first[:200]
        text_for_summary = first
    else:
        text_for_summary = text[:300]

    t = text_for_summary.lower()

    if re.search(r'please (provide|add|include|insert)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 추가/제공 필요."
    if re.search(r'please (update|revise|correct|modify|change)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 수정 필요."
    if re.search(r'please (clarify|explain|confirm)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 명확화 필요."
    if re.search(r'please (remove|delete)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 삭제 필요."
    if re.search(r'please (check|review|verify|ensure)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 확인 필요."
    if re.search(r'(missing|not (included|provided|present))', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 누락. 추가 필요."
    if re.search(r'(incorrect|wrong|error|mistake)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 오류. 수정 필요."
    if re.search(r'(inconsisten)', t):
        subject = _extract_subject(text_for_summary)
        return f"{subject} 일관성 부족. 통일 필요."
    if re.search(r'typo|spelling', t):
        return f"오타 수정 필요."
    if re.search(r'resolution.*(low|poor|insufficient)', t):
        return f"이미지 해상도 부족. 개선 필요."

    subject = _extract_subject(text_for_summary)
    if category == "Technical":
        return f"{subject} 기술적 확인/수정 필요."
    elif category == "FigTable":
        return f"{subject} 그림/표 수정 필요."
    elif category == "Reference":
        return f"{subject} 참조 수정 필요."
    elif category == "Format":
        return f"{subject} 포맷 수정 필요."
    elif category == "Typo":
        return f"오타/표기 수정 필요."
    else:
        return f"{subject} 검토 및 수정 필요."


def _extract_subject(text):
    """Extract the main subject/topic from comment text."""
    text = text.strip()
    text = re.sub(r'^Please\s+', '', text, flags=re.IGNORECASE)
    for sep in ['. ', ', ', ' - ', '\n']:
        if sep in text[:120]:
            text = text[:text.index(sep)]
            break
    if len(text) > 80:
        text = text[:80] + '...'
    return text


def determine_status(status_raw, response_text):
    """Map source status + response to valid CT status."""
    status_lower = str(status_raw).lower().strip() if status_raw else ''
    has_response = bool(response_text and str(response_text).strip())

    if not has_response:
        if 'closed' in status_lower:
            return "Accepted"
        return "Noted"

    resp_lower = str(response_text).lower()
    if any(w in resp_lower for w in ['noted', 'acknowledged', 'we note']):
        return "Noted"
    if any(w in resp_lower for w in ['reject', 'disagree', 'not applicable', 'n/a']):
        return "Rejected"
    if any(w in resp_lower for w in ['applied', 'updated', 'corrected', 'revised', 'uploaded', 'done']):
        return "Accepted"
    if 'closed' in status_lower:
        return "Accepted"
    return "Accepted"


# ============================================================
# EDF Parser
# ============================================================

def parse_edf(filepath):
    """Parse EDF format comment sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = None
    for name in ['Comment sheet', 'Comments', 'Sheet1']:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if name.lower() not in ('cs rules', 'instructions', 'template'):
                ws = wb[name]
                break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    header_row = None
    metadata = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        a_val = str(row[0].value).strip() if row[0].value else ''
        if a_val.lower() == 'item':
            header_row = row_idx
            break
        if 'commented document' in a_val.lower() or 'document' in a_val.lower():
            for c in row:
                if c.value and 'SHR-DEV' in str(c.value):
                    metadata['doc_number'] = str(c.value)
        if 'reception date' in a_val.lower() or 'last issue' in a_val.lower():
            for c in row:
                if c.value and isinstance(c.value, datetime):
                    metadata['date'] = c.value.strftime('%Y-%m-%d')
                elif c.value and re.match(r'\d{1,2}/\d{1,2}/\d{2,4}', str(c.value)):
                    metadata['date'] = str(c.value)

    if header_row is None:
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
            vals = [str(c.value).strip().lower() if c.value else '' for c in row[:5]]
            if 'item' in vals or 'comment no' in vals:
                header_row = row_idx
                break

    if header_row is None:
        print(f"  WARNING: Could not find header row in {filepath}")
        return [], metadata

    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    col_map = {}
    for c in header_cells:
        v = str(c.value).strip().lower() if c.value else ''
        col_idx = c.column - 1
        if v in ('item', 'comment no', 'comment no.', '#', 'number'):
            col_map['num'] = col_idx
        elif v in ('version', 'rev', 'revision'):
            col_map['version'] = col_idx
        elif v in ('page',):
            col_map['page'] = col_idx
        elif v in ('section',):
            col_map['section'] = col_idx
        elif 'comment' in v and 'response' not in v and 'contractor' not in v:
            if 'from' not in v and 'written' not in v:
                col_map['comment'] = col_idx
        elif 'response' in v or 'contractor' in v:
            if 'from' not in v and 'responded' not in v:
                col_map['response'] = col_idx
        elif v in ('status',):
            col_map['status'] = col_idx
        elif 'priority' in v:
            col_map['priority'] = col_idx
        elif 'from' in v and 'written' in v:
            col_map['from_written'] = col_idx
        elif 'from' in v and 'responded' in v:
            col_map['from_responded'] = col_idx

    if 'comment' not in col_map:
        col_map['comment'] = 5
    if 'response' not in col_map:
        col_map['response'] = 7
    if 'status' not in col_map:
        col_map['status'] = 8
    if 'num' not in col_map:
        col_map['num'] = 0

    comments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = list(row)
        num_val = cells[col_map['num']].value if col_map['num'] < len(cells) else None

        if num_val is None:
            continue
        if not isinstance(num_val, (int, float)):
            try:
                num_val = int(num_val)
            except (ValueError, TypeError):
                continue

        comment_text = str(cells[col_map['comment']].value).strip() if col_map['comment'] < len(cells) and cells[col_map['comment']].value else ''
        response_text = str(cells[col_map['response']].value).strip() if col_map['response'] < len(cells) and cells[col_map['response']].value else ''
        status_raw = str(cells[col_map['status']].value).strip() if col_map['status'] < len(cells) and cells[col_map['status']].value else ''
        priority_raw = str(cells[col_map.get('priority', 99)].value).strip() if col_map.get('priority', 99) < len(cells) and cells[col_map.get('priority', 99)].value else ''

        section_parts = []
        if 'section' in col_map and col_map['section'] < len(cells) and cells[col_map['section']].value:
            section_parts.append(str(cells[col_map['section']].value).strip())
        if 'page' in col_map and col_map['page'] < len(cells) and cells[col_map['page']].value:
            section_parts.append(f"p.{cells[col_map['page']].value}")
        section = ', '.join(section_parts)

        reviewer = ''
        if 'from_written' in col_map and col_map['from_written'] < len(cells) and cells[col_map['from_written']].value:
            reviewer = str(cells[col_map['from_written']].value).strip()

        if not comment_text or comment_text == 'None':
            continue

        comments.append({
            'num': int(num_val),
            'section': section,
            'comment_text': comment_text,
            'response_text': response_text if response_text != 'None' else '',
            'status_raw': status_raw,
            'priority_raw': priority_raw,
            'reviewer': reviewer,
        })

    return comments, metadata


# ============================================================
# Vena Parser
# ============================================================

def parse_vena(filepath):
    """Parse Vena Energy format comment sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    metadata = {}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for c in row:
            if c.value and 'Project Title' in str(c.value):
                for c2 in row:
                    if c2.column > c.column and c2.value:
                        metadata['project_name'] = str(c2.value).strip()
                        break
            if c.value and 'CRS Number' in str(c.value):
                for c2 in row:
                    if c2.column > c.column and c2.value:
                        metadata['crs_number'] = str(c2.value).strip()
                        break
            if c.value and 'Revision' in str(c.value):
                for c2 in row:
                    if c2.column > c.column and c2.value:
                        metadata['revision'] = str(c2.value).strip()
                        break

    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        vals = [str(c.value).strip().lower() if c.value else '' for c in row[:5]]
        if 'number' in vals or 'item' in vals:
            header_row = row_idx
            break

    if header_row is None:
        print(f"  WARNING: Could not find header row in {filepath}")
        return [], metadata

    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    col_map = {}
    for c in header_cells:
        v = str(c.value).strip().lower() if c.value else ''
        idx = c.column - 1
        if v in ('number', 'item', '#'):
            col_map['num'] = idx
        elif v == 'date':
            col_map['date'] = idx
        elif 'comment by' in v or v == 'from':
            col_map['reviewer'] = idx
        elif v == 'discipline':
            col_map['discipline'] = idx
        elif v == 'deliverable':
            col_map['deliverable'] = idx
        elif 'comment detail' in v or v == 'comments':
            col_map['comment'] = idx
        elif v == 'response':
            col_map['response'] = idx
        elif 'comment (2nd)' in v or 'comment 2' in v:
            col_map['comment_2nd'] = idx
        elif 'closed' in v or 'status' in v:
            col_map['status'] = idx

    if 'comment' not in col_map:
        col_map['comment'] = 6
    if 'response' not in col_map:
        col_map['response'] = 7
    if 'num' not in col_map:
        col_map['num'] = 0

    comments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = list(row)
        num_val = cells[col_map['num']].value if col_map['num'] < len(cells) else None

        if num_val is None:
            continue
        if not isinstance(num_val, (int, float)):
            try:
                num_val = int(num_val)
            except (ValueError, TypeError):
                continue

        comment_text = str(cells[col_map['comment']].value).strip() if col_map['comment'] < len(cells) and cells[col_map['comment']].value else ''
        response_text = str(cells[col_map['response']].value).strip() if col_map['response'] < len(cells) and cells[col_map['response']].value else ''
        status_raw = ''
        if 'status' in col_map and col_map['status'] < len(cells) and cells[col_map['status']].value:
            status_raw = str(cells[col_map['status']].value).strip()

        section_parts = []
        if 'discipline' in col_map and col_map['discipline'] < len(cells) and cells[col_map['discipline']].value:
            section_parts.append(str(cells[col_map['discipline']].value).strip())
        if 'deliverable' in col_map and col_map['deliverable'] < len(cells) and cells[col_map['deliverable']].value:
            section_parts.append(str(cells[col_map['deliverable']].value).strip())
        section = ' - '.join(section_parts)

        reviewer = ''
        if 'reviewer' in col_map and col_map['reviewer'] < len(cells) and cells[col_map['reviewer']].value:
            reviewer = str(cells[col_map['reviewer']].value).strip()

        if 'comment_2nd' in col_map and col_map['comment_2nd'] < len(cells) and cells[col_map['comment_2nd']].value:
            c2 = str(cells[col_map['comment_2nd']].value).strip()
            if c2 and c2 != 'None':
                comment_text += f"\n[Follow-up] {c2}"

        if not comment_text or comment_text == 'None':
            continue

        comments.append({
            'num': int(num_val),
            'section': section,
            'comment_text': comment_text,
            'response_text': response_text if response_text != 'None' else '',
            'status_raw': status_raw,
            'priority_raw': '',
            'reviewer': reviewer,
        })

    return comments, metadata


# ============================================================
# Vena Korean Translation Merger
# ============================================================

def load_vena_korean(filepath):
    """Load Korean translations from Vena file. Returns dict {comment_num: korean_text}."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        print(f"  WARNING: Could not load Korean file {filepath}: {e}")
        return {}

    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        vals = [str(c.value).strip().lower() if c.value else '' for c in row[:5]]
        if 'number' in vals or 'item' in vals:
            header_row = row_idx
            break

    if header_row is None:
        return {}

    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    comment_col = None
    num_col = 0
    for c in header_cells:
        v = str(c.value).strip().lower() if c.value else ''
        if v in ('number', 'item', '#'):
            num_col = c.column - 1
        if 'comment' in v and 'response' not in v and 'by' not in v and '2nd' not in v:
            comment_col = c.column - 1

    if comment_col is None:
        comment_col = 6

    korean_map = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = list(row)
        num_val = cells[num_col].value if num_col < len(cells) else None
        if num_val is None or not isinstance(num_val, (int, float)):
            continue
        text = cells[comment_col].value if comment_col < len(cells) else None
        if text:
            korean_map[int(num_val)] = str(text).strip()

    return korean_map


# ============================================================
# Orsted Parser
# ============================================================

ABC_PATTERN = re.compile(r'^([a-c])\)\s*(.*)', re.DOTALL)
NUM_DOT_PATTERN = re.compile(r'^(\d+)\.?\s*$')

def _strip_abc_prefix(text):
    """Remove a)/b)/c) prefix from text and return (sub_letter, cleaned_text)."""
    if not text:
        return None, ''
    text = text.strip()
    m = ABC_PATTERN.match(text)
    if m:
        return m.group(1), m.group(2).strip()
    return None, text


def parse_orsted(filepath):
    """Parse Orsted format comment sheet.
    Format: Col A has 'NN.' repeated 3 rows per comment,
    Col E has 'a)/b)/c)' prefixed text, Col H has replies.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the comment sheet
    ws = None
    for name in wb.sheetnames:
        if 'comment' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Extract metadata from header area (rows 1-12)
    metadata = {}
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=False):
        cells = list(row)
        a_val = str(cells[0].value).strip() if cells[0].value else ''
        if 'project title' in a_val.lower():
            for c in cells[3:]:
                if c.value and str(c.value).strip():
                    metadata['project_title'] = str(c.value).strip()
                    break
        elif 'document title' in a_val.lower():
            for c in cells[3:]:
                if c.value and str(c.value).strip():
                    metadata['doc_title'] = str(c.value).strip()
                    break
        elif 'contractor name' in a_val.lower():
            for c in cells[3:]:
                if c.value and str(c.value).strip():
                    metadata['contractor'] = str(c.value).strip()
                    break

    # Find reviewer from row 2 (Comments approved by)
    try:
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=False))[0]
        for c in row2:
            v = str(c.value).strip() if c.value else ''
            if len(v) >= 3 and len(v) <= 10 and v.isalpha() and v.isupper():
                metadata['reviewer'] = v
                break
    except Exception:
        pass

    # Find header row (contains 'Comment no.')
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        for c in row:
            v = str(c.value).strip().lower() if c.value else ''
            if 'comment no' in v:
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        print(f"  WARNING: Could not find header row in {filepath}")
        return [], metadata

    # Read all data rows and group by comment number
    ACKNOWLEDGMENTS = ['thank you', 'understood', 'comment can be closed', 'looks good',
                       'no further', 'agreed', 'ok', 'noted', 'well noted']
    ERROR_VALUES = {'#VALUE!', '#REF!', '#N/A', '#NAME?', '#DIV/0!', '#NULL!'}

    groups = {}  # {comment_num: {'a': {...}, 'b': {...}, 'c': {...}}}

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        cells = list(row)
        a_val = cells[0].value
        if a_val is None:
            continue

        # Parse comment number from col A: "01.", "02.", etc.
        a_str = str(a_val).strip()
        num_match = NUM_DOT_PATTERN.match(a_str)
        if not num_match:
            continue
        num = int(num_match.group(1))

        # Parse sub-letter from col E prefix
        e_raw = str(cells[4].value).strip() if len(cells) > 4 and cells[4].value else ''
        if e_raw in ERROR_VALUES:
            e_raw = ''
        sub, e_text = _strip_abc_prefix(e_raw)
        if sub is None:
            continue  # Not a valid a/b/c row

        # Parse reply from col H
        h_raw = str(cells[7].value).strip() if len(cells) > 7 and cells[7].value else ''
        if h_raw in ERROR_VALUES:
            h_raw = ''
        _, h_text = _strip_abc_prefix(h_raw)

        # Other columns
        section = str(cells[3].value).strip() if len(cells) > 3 and cells[3].value else ''
        page = str(cells[2].value).strip() if len(cells) > 2 and cells[2].value else ''
        reviewer = str(cells[5].value).strip() if len(cells) > 5 and cells[5].value else ''
        date_val = str(cells[6].value).strip() if len(cells) > 6 and cells[6].value else ''
        status = str(cells[12].value).strip() if len(cells) > 12 and cells[12].value else ''

        if num not in groups:
            groups[num] = {}
        groups[num][sub] = {
            'text': e_text,
            'reply': h_text,
            'section': section,
            'page': page,
            'reviewer': reviewer,
            'date': date_val,
            'status': status,
        }

    # Build comments from groups
    comments = []
    for num in sorted(groups.keys()):
        g = groups[num]
        a_data = g.get('a', {})
        a_text = a_data.get('text', '')

        # Skip empty placeholder comments (no a) text)
        if not a_text:
            continue

        # Section with page
        section = a_data.get('section', '')
        page = a_data.get('page', '')
        if page and page != 'None':
            section = f"p.{page}, {section}" if section else f"p.{page}"

        # Main reply from a)
        reply = a_data.get('reply', '')

        # Collect b/c follow-ups
        followups = []
        followup_replies = []
        for sub in ['b', 'c']:
            if sub in g:
                sub_text = g[sub].get('text', '')
                sub_reply = g[sub].get('reply', '')
                if sub_text and not any(ack in sub_text.lower() for ack in ACKNOWLEDGMENTS):
                    followups.append(f"[{sub}] {sub_text}")
                if sub_reply:
                    followup_replies.append(f"[{sub}] {sub_reply}")

        # Build full comment text
        comment_text = a_text
        if followups:
            comment_text += '\n' + '\n'.join(followups)

        # Build full reply text
        response_text = reply
        if followup_replies:
            response_text += '\n' + '\n'.join(followup_replies)

        # Reviewer: prefer a) row, fallback to metadata
        reviewer = a_data.get('reviewer', '') or metadata.get('reviewer', '')

        comments.append({
            'num': num,
            'section': section,
            'comment_text': comment_text,
            'response_text': response_text,
            'status_raw': a_data.get('status', ''),
            'priority_raw': '',
            'reviewer': reviewer,
        })

    return comments, metadata


# ============================================================
# JSON Generator
# ============================================================

def generate_ct_json(comments, project_meta, batch_meta, source_file, korean_map=None):
    """Generate CT JSON from parsed comments."""
    ct_comments = []

    for i, c in enumerate(comments, 1):
        severity, category, confidence, excluded, exclude_reason = classify_comment(
            c['comment_text'], c.get('priority_raw')
        )
        status = determine_status(c['status_raw'], c['response_text'])
        tags = extract_tags(c['comment_text'], c['num'])

        if korean_map and c['num'] in korean_map:
            summary_ko = korean_map[c['num']]
        else:
            summary_ko = generate_summary_ko(c['comment_text'], category)

        ct_comments.append({
            "comment_number": i,
            "section": c['section'],
            "comment_text": c['comment_text'],
            "summary_ko": summary_ko,
            "severity": severity,
            "category": category,
            "status": status,
            "response_text": c['response_text'],
            "assignee": batch_meta.get('assignee', 'KJH'),
            "excluded": excluded,
            "exclude_reason": exclude_reason,
            "confidence": confidence,
            "tags": tags,
        })

    output = {
        "project": project_meta,
        "batch": {
            **batch_meta,
            "source_file": source_file,
        },
        "comments": ct_comments,
    }

    return output


def save_json(data, output_dir, project_code, revision, comment_type):
    """Save CT JSON file."""
    filename = f"CT_{project_code}_{revision}_{comment_type}.json"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return filepath


def print_summary(data, filepath):
    """Print summary of generated JSON."""
    comments = data['comments']
    total = len(comments)
    if total == 0:
        print("No comments found.")
        return
    major = sum(1 for c in comments if c['severity'] == 'Major')
    minor = sum(1 for c in comments if c['severity'] == 'Minor')
    excluded = sum(1 for c in comments if c['excluded'])
    typo = sum(1 for c in comments if c['category'] == 'Typo' and not c['excluded'])
    readability = sum(1 for c in comments if c['category'] == 'Readability' and not c['excluded'])
    figtable = sum(1 for c in comments if c['category'] == 'FigTable' and not c['excluded'])
    fmt = sum(1 for c in comments if c['category'] == 'Format' and not c['excluded'])
    ref = sum(1 for c in comments if c['category'] == 'Reference' and not c['excluded'])
    technical = sum(1 for c in comments if c['category'] == 'Technical' and not c['excluded'])
    low_conf = sum(1 for c in comments if c['confidence'] == 'Low')
    ko_filled = sum(1 for c in comments if c['summary_ko'])

    print(f"""
=== Comment Tracker Import Summary ===
File: {data['batch']['source_file']}
Output: {os.path.basename(filepath)}
Project: {data['project']['project_code']} | Rev: {data['batch']['revision']} | Type: {data['batch']['comment_type']}
Total comments: {total} items
  Major (Technical): {technical} ({technical/total*100:.0f}%)
  Minor (Document): {minor-excluded} ({(minor-excluded)/total*100:.0f}%)
    Typo:        {typo}
    Readability: {readability}
    FigTable:    {figtable}
    Format:      {fmt}
    Reference:   {ref}
Excluded: {excluded} items
Korean summaries: {ko_filled}/{total} filled
Low-confidence items: {low_conf} (review needed)
""")


if __name__ == '__main__':
    print("CT Parser loaded. Use functions directly or import.")
